"""
audit_leads.py — Auditoría diaria: Google Sheet de descargas Meta vs Sperant.

Flujo:
  1. Descarga el Google Sheet "leads <mes> <año>" como xlsx (export público, sin auth).
  2. Lee cada hoja (1 por proyecto) y extrae email / teléfono / DNI / nombre.
  3. Matchea cada lead contra TODO tuna.clientes (Redshift) con cascada:
       email exacto (+ normalización de typos) → email parcial → tel (últimos 9) →
       DNI → nombre fuzzy.
  4. Clasifica cada lead:
       NUEVO_OK        — primer touchpoint del cliente en mayo en ESE proyecto
       RECAP_HIST      — ya estaba en el mismo proyecto en un mes anterior
       RECAP_CROSS     — ya estaba en OTRO proyecto del CRM
       RECAP_FUERA_PER — está en CRM pero su actividad no es de este mes
       TEST            — lead de prueba (tel 7777..., email noexisto, etc.)
       FALTANTE_REAL   — no existe en tuna.clientes por NINGÚN criterio
  5. Cuenta manuales creados en Sperant en el mes (origen=manual, creación de cliente)
     por proyecto → mide salud del proceso de reconciliación.
  6. Imprime reporte. Exit code 2 si hay FALTANTE_REAL (para que el workflow alerte).

Config por env:
  REDSHIFT_HOST / PORT / DB / USER / PASSWORD  (ya en GitHub secrets del repo)
  LEADS_SHEET_ID   — ID del Google Sheet del mes (repo variable, editable sin tocar código)
  AUDIT_YEAR / AUDIT_MONTH — opcional, default = mes actual

Mapeo hoja → código Sperant: editar SHEET2CODIGO abajo cuando se agreguen proyectos.
"""
import os
import re
import sys
import json
import unicodedata
from datetime import datetime, timezone

import requests
import psycopg2
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

SHEET_ID = os.environ.get("LEADS_SHEET_ID", "").strip() or \
    "151SxHkCYPZFkx7V4nIlv9ZAq5VYcbvlnsp_XrJfLNFE"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# Hoja del Sheet → código_proyecto en tuna.interacciones.
# vartel / seres no están en el ETL de Sperant → se omiten (informativo).
SHEET2CODIGO = {
    "strena": "STRN",
    "monte alegre": "MA",
    "gemma": "GEMMA",
    "melgar": "MELGAR",
    "palacios": "PALACIOS",
    "romaña": "R125",
}
SHEETS_INFORMATIVAS = {"vartel", "seres"}  # se reportan pero no se auditan vs Sperant

now = datetime.now(timezone.utc)


def _env_int(key, default):
    """os.environ.get pero tolera string vacío (workflow_dispatch sin input)."""
    v = os.environ.get(key, "")
    return int(v) if v.strip() else int(default)


AUDIT_YEAR = _env_int("AUDIT_YEAR", now.year)
AUDIT_MONTH = _env_int("AUDIT_MONTH", now.month)

EMAIL_TYPOS = [
    ("gmal.com", "gmail.com"), ("gnail.com", "gmail.com"),
    ("gmail.commail.com", "gmail.com"), ("gamail.com", "gmail.com"),
    ("hotmail.con", "hotmail.com"), ("hotmail.co", "hotmail.com"),
    ("hotmal.com", "hotmail.com"), ("hotamil.com", "hotmail.com"),
    ("hotmat.com", "hotmail.com"), ("yahoo.con", "yahoo.com"),
    ("outlook.con", "outlook.com"), ("icloud.con", "icloud.com"),
    ("gmail.con", "gmail.com"),
]

TEST_PATTERNS = re.compile(
    r"(1234567\d*|111111111|2222222|7777777|9999999|0000000|"
    r"noexisto|prueba|test@|@test|moars99)"
)


# ─────────────────────────────────────────────────────────────────────────────
# Normalizadores
# ─────────────────────────────────────────────────────────────────────────────

def norm_email(s):
    if s is None:
        return None
    s = str(s).strip().lower().replace(" ", "")
    if "@" not in s or "." not in s:
        return None
    for typo, fix in EMAIL_TYPOS:
        if s.endswith(typo):
            s = s[: -len(typo)] + fix
            break
    return s


def norm_tel(s):
    if s is None:
        return None
    d = re.sub(r"\D", "", str(s))
    if not d:
        return None
    if d.startswith("51") and len(d) >= 11:
        d = d[2:]
    return d[-9:] if len(d) >= 9 else None


def norm_dni(s):
    if s is None:
        return None
    d = re.sub(r"\D", "", str(s))
    return d if 7 <= len(d) <= 8 else None


def strip_accents(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


SKIP_NAME_KW = [
    "flats_", "penthouse", "m2", "duplex", "_meses", "este_mes", "aún_no",
    "quiero_", "sí,", "_piso", "no contesta", "desestimado", "bajo",
    "intermedio", "alto", "fuera_", "por_", "compr", "busca", "fecha_",
    "solicit", "solo_", "precio", "$", "suma total", "counta",
]


def extract_leads(ws):
    leads = []
    for row in ws.iter_rows(values_only=True):
        email = None
        for v in row:
            e = norm_email(v)
            if e:
                email = e
                break
        if not email:
            continue

        tel = dni = name = None
        for v in row:
            if v is None:
                continue
            sv = str(v).strip()
            if "@" in sv:
                continue
            digits = re.sub(r"\D", "", sv)
            if tel is None and len(digits) >= 9 and (
                sv.startswith("51") or sv.startswith("9") or len(digits) == 9
            ):
                if "/" not in sv or len(sv) > 10:  # evitar fechas
                    tel = norm_tel(v)
            if dni is None and sv.isdigit() and 7 <= len(sv) <= 8:
                dni = sv

        for v in row:
            if v is None:
                continue
            sv = str(v).strip()
            if "@" in sv or not sv:
                continue
            if re.match(r"^[\d\s/_\-:\.]+$", sv):
                continue
            sl = sv.lower()
            if any(k in sl for k in SKIP_NAME_KW):
                continue
            if " " in sv and 4 < len(sv) < 80:
                name = sv
                break

        leads.append({"email": email, "tel": tel, "dni": dni, "name": name})
    return leads


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print(f"=== AUDITORÍA LEADS SHEET vs SPERANT — {AUDIT_YEAR}-{AUDIT_MONTH:02d} ===")
    print(f"Sheet ID: {SHEET_ID}")

    # 1) Descargar xlsx
    r = requests.get(EXPORT_URL, timeout=60)
    r.raise_for_status()
    with open("/tmp/_leads.xlsx", "wb") as f:
        f.write(r.content)
    wb = load_workbook("/tmp/_leads.xlsx", data_only=True)
    print(f"Hojas: {wb.sheetnames}\n")

    sheet_leads = {}
    for sn in wb.sheetnames:
        leads = extract_leads(wb[sn])
        sheet_leads[sn] = leads
        tag = "" if sn in SHEET2CODIGO else " (informativa, no se audita)"
        print(f"  {sn:15s}: {len(leads):4d} leads{tag}")

    # 2) Redshift
    conn = psycopg2.connect(
        host=os.environ["REDSHIFT_HOST"],
        port=int(os.environ["REDSHIFT_PORT"]),
        dbname=os.environ["REDSHIFT_DB"],
        user=os.environ["REDSHIFT_USER"],
        password=os.environ["REDSHIFT_PASSWORD"],
        connect_timeout=30,
        sslmode="require",
    )
    cur = conn.cursor()

    # Universo de clientes
    print("\nCargando universo tuna.clientes ...")
    cur.execute("""
        SELECT id, LOWER(TRIM(email)), telefono, celulares, documento,
               LOWER(TRIM(nombres || ' ' || COALESCE(apellidos, '')))
        FROM tuna.clientes
    """)
    idx_email, idx_tel, idx_dni = {}, {}, {}
    cli_nombre = {}
    for cid, email, tel, cel, doc, nombre in cur.fetchall():
        ne = norm_email(email)
        if ne:
            idx_email.setdefault(ne, []).append(cid)
        for x in (tel, cel):
            nt = norm_tel(x)
            if nt:
                idx_tel.setdefault(nt, []).append(cid)
        nd = norm_dni(doc)
        if nd:
            idx_dni.setdefault(nd, []).append(cid)
        if nombre:
            cli_nombre[cid] = strip_accents(nombre)
    print(f"  email={len(idx_email)}  tel={len(idx_tel)}  dni={len(idx_dni)}")

    # Primera interacción por cliente×proyecto
    cur.execute("""
        SELECT cliente_id, codigo_proyecto, MIN(fecha_creacion)
        FROM tuna.interacciones
        WHERE codigo_proyecto IS NOT NULL
        GROUP BY cliente_id, codigo_proyecto
    """)
    cli_inter = {}
    for cid, proj, fecha in cur.fetchall():
        cli_inter.setdefault(cid, {})[proj] = fecha

    # Manuales del mes por proyecto
    cur.execute(f"""
        SELECT codigo_proyecto, COUNT(DISTINCT cliente_id)
        FROM tuna.interacciones
        WHERE origen = 'manual' AND tipo_interaccion = 'creación de cliente'
          AND DATE_PART('year', fecha_creacion) = {AUDIT_YEAR}
          AND DATE_PART('month', fecha_creacion) = {AUDIT_MONTH}
        GROUP BY codigo_proyecto
    """)
    manuales = {r[0]: r[1] for r in cur.fetchall()}

    # Clientes únicos con interacción en el mes por proyecto (universo Sperant)
    cur.execute(f"""
        SELECT codigo_proyecto, COUNT(DISTINCT cliente_id)
        FROM tuna.interacciones
        WHERE DATE_PART('year', fecha_creacion) = {AUDIT_YEAR}
          AND DATE_PART('month', fecha_creacion) = {AUDIT_MONTH}
        GROUP BY codigo_proyecto
    """)
    clientes_mes = {r[0]: r[1] for r in cur.fetchall()}

    # 3) Clasificar
    def fuzzy_name(name):
        if not name:
            return False
        partes = [p for p in strip_accents(name.lower()).split() if len(p) >= 3]
        if len(partes) < 2:
            return False
        primer, ultimo = partes[0], partes[-1]
        for nom in cli_nombre.values():
            if primer in nom and ultimo in nom:
                return True
        return False

    resumen = []
    faltantes = []
    print(f"\n{'='*92}")
    print("RESULTADO POR PROYECTO")
    print(f"{'='*92}")

    for tab, codigo in SHEET2CODIGO.items():
        leads = sheet_leads.get(tab, [])
        c = {"NUEVO_OK": 0, "RECAP_HIST": 0, "RECAP_CROSS": 0,
             "RECAP_FUERA_PER": 0, "TEST": 0, "FALTANTE_REAL": 0}

        for L in leads:
            blob = f"{L.get('email') or ''} {L.get('tel') or ''}"
            if TEST_PATTERNS.search(blob):
                c["TEST"] += 1
                continue

            cids = set()
            if L["email"] and L["email"] in idx_email:
                cids.update(idx_email[L["email"]])
            if not cids and L["email"] and "@" in L["email"]:
                local = L["email"].split("@")[0]
                if len(local) >= 6:
                    cur.execute(
                        "SELECT id FROM tuna.clientes "
                        "WHERE LOWER(email) LIKE %s LIMIT 1",
                        (local + "@%",),
                    )
                    rr = cur.fetchone()
                    if rr:
                        cids.add(rr[0])
            if not cids and L["tel"] and L["tel"] in idx_tel:
                cids.update(idx_tel[L["tel"]])
            if not cids and L["dni"] and L["dni"] in idx_dni:
                cids.update(idx_dni[L["dni"]])
            if not cids and fuzzy_name(L.get("name")):
                cids.add("_fuzzy_")

            if not cids:
                c["FALTANTE_REAL"] += 1
                faltantes.append({**L, "tab": tab, "codigo": codigo})
                continue

            clase = "RECAP_FUERA_PER"
            for cid in cids:
                if cid == "_fuzzy_":
                    clase = "RECAP_HIST"
                    break
                ints = cli_inter.get(cid, {})
                pe = ints.get(codigo)
                if pe and pe.year == AUDIT_YEAR and pe.month == AUDIT_MONTH:
                    clase = "NUEVO_OK"
                    break
                if any(f.year == AUDIT_YEAR and f.month == AUDIT_MONTH
                       for f in ints.values()):
                    clase = "RECAP_CROSS"
                elif codigo in ints:
                    clase = "RECAP_HIST"
            c[clase] += 1

        man = manuales.get(codigo, 0)
        salud = "OK" if c["FALTANTE_REAL"] == 0 else (
            "REVISAR" if man >= c["FALTANTE_REAL"] else "GAP OPERATIVO"
        )
        print(
            f"\n  {tab.upper()} ({codigo}) — {len(leads)} leads Sheet | "
            f"manuales creados este mes: {man}  → {salud}"
        )
        for k, v in c.items():
            if v:
                print(f"      {k:16s}: {v}")
        resumen.append({"tab": tab, "codigo": codigo, "total": len(leads),
                         "manuales": man, "salud": salud,
                         "sperant_mes": clientes_mes.get(codigo, 0), **c})

    # Resumen global
    print(f"\n{'='*92}\nRESUMEN GLOBAL\n{'='*92}")
    tot = {"total": 0, "NUEVO_OK": 0, "RECAP_HIST": 0, "RECAP_CROSS": 0,
           "RECAP_FUERA_PER": 0, "TEST": 0, "FALTANTE_REAL": 0}
    hdr = (f"  {'Proyecto':14s}{'Tot':>5s}{'Nuevo':>7s}{'Recap':>7s}"
           f"{'Cross':>7s}{'Test':>6s}{'FALTA':>7s}{'Manual':>8s}  Salud")
    print(hdr)
    for r in resumen:
        print(f"  {r['tab']:14s}{r['total']:>5d}{r['NUEVO_OK']:>7d}"
              f"{r['RECAP_HIST']:>7d}{r['RECAP_CROSS']:>7d}{r['TEST']:>6d}"
              f"{r['FALTANTE_REAL']:>7d}{r['manuales']:>8d}  {r['salud']}")
        for k in tot:
            tot[k] += r[k if k != "total" else "total"]
    print(f"  {'TOTAL':14s}{tot['total']:>5d}{tot['NUEVO_OK']:>7d}"
          f"{tot['RECAP_HIST']:>7d}{tot['RECAP_CROSS']:>7d}{tot['TEST']:>6d}"
          f"{tot['FALTANTE_REAL']:>7d}")

    # Detalle faltantes
    if faltantes:
        print(f"\n{'='*92}\nFALTANTES REALES ({len(faltantes)}) — no existen en Sperant\n{'='*92}")
        for L in faltantes:
            print(f"  · {L['codigo']:9s} {(L.get('name') or '?')[:32]:32s} "
                  f"{L.get('email') or '?':38s} tel={L.get('tel') or '-'}")

    # ── Movimiento vs día anterior (evidencia si Alonso concilia a diario) ──
    # Snapshot diario versionado en el repo: audit/history/<periodo>/<fecha>.json
    hist_dir = os.path.join(
        os.path.dirname(__file__), "history", f"{AUDIT_YEAR}-{AUDIT_MONTH:02d}"
    )
    os.makedirs(hist_dir, exist_ok=True)
    hoy_str = now.strftime("%Y-%m-%d")

    # Snapshot de hoy: por proyecto {sheet, sperant_mes, manuales, faltantes}
    snap_hoy = {
        r["codigo"]: {
            "sheet": r["total"],
            "sperant_mes": r["sperant_mes"],
            "manuales": r["manuales"],
            "faltantes": r["FALTANTE_REAL"],
        }
        for r in resumen
    }

    # Buscar el snapshot anterior más reciente (≠ hoy) del mismo periodo
    prev = None
    prev_fecha = None
    try:
        files = sorted(
            f for f in os.listdir(hist_dir)
            if f.endswith(".json") and f[:-5] < hoy_str
        )
        if files:
            prev_fecha = files[-1][:-5]
            with open(os.path.join(hist_dir, files[-1])) as fh:
                prev = json.load(fh)
    except FileNotFoundError:
        pass

    print(f"\n{'='*92}")
    if prev:
        print(f"MOVIMIENTO vs {prev_fecha} (¿Alonso está conciliando a diario?)")
        print(f"{'='*92}")
        print(f"  {'Proyecto':12s}{'ΔSheet':>9s}{'ΔSperant':>10s}"
              f"{'ΔManual':>9s}{'ΔFaltan':>9s}  Lectura")
        for r in resumen:
            cod = r["codigo"]
            p = prev.get(cod, {})
            d_sheet = r["total"] - p.get("sheet", r["total"])
            d_sper = r["sperant_mes"] - p.get("sperant_mes", r["sperant_mes"])
            d_man = r["manuales"] - p.get("manuales", r["manuales"])
            d_falt = r["FALTANTE_REAL"] - p.get("faltantes", r["FALTANTE_REAL"])
            # Lectura: Alonso agregó al Sheet pero el equipo no creó manuales
            if d_sheet > 0 and d_man == 0 and d_falt >= 0:
                lectura = "⚠ Sheet creció, 0 manuales"
            elif d_sheet == 0 and prev:
                lectura = "Sheet sin cambios"
            else:
                lectura = "ok"
            print(f"  {r['tab']:12s}{d_sheet:>+9d}{d_sper:>+10d}"
                  f"{d_man:>+9d}{d_falt:>+9d}  {lectura}")
        print("\n  ΔSheet  = leads que Alonso agregó al Excel desde el snapshot anterior")
        print("  ΔSperant= clientes nuevos que entraron al CRM (sync auto + manual)")
        print("  ΔManual = leads que el equipo reconcilió a mano")
        print("  ΔFaltan = variación de faltantes (sube = peor)")
    else:
        print("MOVIMIENTO vs día anterior")
        print(f"{'='*92}")
        print("  (primer snapshot del periodo — sin comparación aún. "
              "Mañana ya habrá deltas.)")

    out = {
        "fecha": now.isoformat(),
        "periodo": f"{AUDIT_YEAR}-{AUDIT_MONTH:02d}",
        "comparado_vs": prev_fecha,
        "resumen": resumen,
        "totales": tot,
        "faltantes": faltantes,
    }
    with open("/tmp/audit_result.json", "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)

    # Guardar snapshot de hoy (lo commitea el workflow)
    with open(os.path.join(hist_dir, f"{hoy_str}.json"), "w") as f:
        json.dump(snap_hoy, f, indent=2, ensure_ascii=False)

    conn.close()

    # Exit code: 2 si hay faltantes reales o algún proyecto en GAP OPERATIVO
    gap = any(r["salud"] == "GAP OPERATIVO" for r in resumen)
    if tot["FALTANTE_REAL"] > 0 or gap:
        print(f"\n⚠ {tot['FALTANTE_REAL']} faltantes reales / gap operativo detectado.")
        sys.exit(2)
    print("\n✓ Sin faltantes reales. Conciliación OK.")


if __name__ == "__main__":
    main()
